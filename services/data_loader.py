from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook


# -----------------------------
# Путь к Excel-файлу
# -----------------------------

# Берём путь не относительно текущей папки запуска,
# а относительно самого файла data_loader.py.
#
# Это надёжнее:
# программа сможет найти Excel даже если запуск идёт
# не из корня проекта, а из IDE или другого места.
BASE_DIR = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = BASE_DIR / 'data' / 'khv_courts.xlsx'


# -----------------------------
# Внутренние helper-функции
# -----------------------------

# Проверяет, существует ли Excel-файл по ожидаемому пути.
def ensure_workbook_exists() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            f'Не найден файл Excel: {WORKBOOK_PATH}'
        )


# Преобразует строку Excel в словарь вида:
# {заголовок_колонки: значение}
#
# headers — список заголовков из первой строки листа
# values — значения текущей строки
def build_row_dict(headers: list, values: tuple) -> dict:
    row_dict = {}

    for index, header in enumerate(headers):
        # Если заголовок пустой, такую колонку пропускаем.
        if header is None:
            continue

        row_dict[str(header).strip()] = values[index]

    return row_dict


# Проверяет, является ли строка полностью пустой.
# Такие строки в итоговый список не добавляем.
def is_empty_row(values: tuple) -> bool:
    for value in values:
        if value not in (None, ''):
            return False

    return True


# -----------------------------
# Чтение листов Excel
# -----------------------------

# Загружает все строки листа в виде списка словарей.
#
# Пример результата:
# [
#     {'Категория': 'село', 'Адрес / объект': 'Мирное', ...},
#     {'Категория': 'улица', 'Адрес / объект': 'Ленина', ...},
# ]
#
# Кэшируем результат, чтобы Excel-файл не открывался повторно
# при каждом новом поиске подсудности.
@lru_cache(maxsize=None)
def load_sheet_rows(sheet_name: str) -> list[dict]:
    ensure_workbook_exists()

    workbook = load_workbook(
        WORKBOOK_PATH,
        data_only=True,
        read_only=True
    )

    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f'Лист "{sheet_name}" не найден в файле {WORKBOOK_PATH.name}')

    sheet = workbook[sheet_name]

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        workbook.close()
        return []

    headers = list(rows[0])
    result = []

    for values in rows[1:]:
        if is_empty_row(values):
            continue

        row_dict = build_row_dict(headers, values)
        result.append(row_dict)

    workbook.close()
    return result


# -----------------------------
# Индекс по листу "Участки"
# -----------------------------

# Собирает словарь по номеру участка.
#
# Пример:
# {
#     66: {
#         'Участок': 66,
#         'Районный суд': '...',
#         'Адрес участка': '...',
#         ...
#     },
#     67: {...}
# }
#
# Это удобно, потому что в правилах часто хранится только номер участка,
# а полные контакты подтягиваются уже отсюда.
@lru_cache(maxsize=1)
def get_sections_index() -> dict:
    rows = load_sheet_rows('Участки')
    sections_index = {}

    for row in rows:
        section_number = row.get('Участок')

        # Если номер участка не указан, такую строку пропускаем.
        if section_number in (None, ''):
            continue

        sections_index[section_number] = row

    return sections_index